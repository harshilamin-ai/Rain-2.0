"""
test_rain_api.py
─────────────────────────────────────────────────────────────────────────────
Tests the Rain 2.0 Matchmaker API (POST /match) against every user in
inputdata.xlsx and compares the API's top-3 results to the expected matches
already stored in the sheet.

Usage
─────
  # With a locally running API (default):
  python test_rain_api.py

  # Against a remote or different port:
  python test_rain_api.py --api-url http://your-server:8000

  # Use a specific input file:
  python test_rain_api.py --input path/to/inputdata.xlsx

  # Save output to a specific location:
  python test_rain_api.py --output path/to/results.xlsx

Requirements
────────────
  pip install requests pandas openpyxl

How it works
────────────
  1. Loads inputdata.xlsx — each row is a user.
  2. Parses each user's profile JSON into a UserProfileInfo payload and their
     objective text into a UserObjective payload.
  3. Treats every OTHER user in the file as a NetworkProfile candidate.
  4. Calls POST /match with the assembled MatchRequest.
  5. Takes the API's top-3 returned profile_ids as the predicted matches.
  6. Compares predicted vs expected (match target 1/2/3 columns).
  7. Writes a colour-coded Excel report showing per-rank match/mismatch and
     an overall accuracy summary.

Objective → UserObjective parsing
──────────────────────────────────
The userobjective column is semicolon-delimited sentences, e.g.:
  "Secure enterprise leads … ; Identify one trading desk … ; Find an ops leader …"

Mapping:
  primary_goal      ← first sentence
  secondary_goals   ← remaining sentences
  target_profiles   ← inferred from title keywords in the objective text
  success_signals   ← every sentence (the whole objective drives retrieval)

Profile → NetworkProfile parsing
─────────────────────────────────
The userprofileinfo column is a JSON blob matching UserProfileInfo.
For NetworkProfile we extract:
  profile_id  ← userID
  name        ← userName
  title       ← current_role.title
  company     ← current_role.company
  industry    ← inferred from solutions_offered / career_highlights
  skills      ← [sk.skill for sk in top_skills]
  summary     ← solutions_offered joined
"""

import argparse
import json
import sys
import time
import re
from typing import List, Dict, Optional, Tuple

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CLI args ──────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser(description="Test Rain 2.0 Match API")
parser.add_argument("--api-url",  default="http://localhost:8000",
                    help="Base URL of the running Rain API (default: http://localhost:8000)")
parser.add_argument("--input",    default="test_dataset.xlsx",
                    help="Path to test_dataset.xlsx (default: test_dataset.xlsx)")
parser.add_argument("--output",   default="api_comparison_results.xlsx",
                    help="Path to write comparison Excel (default: api_comparison_results.xlsx)")
parser.add_argument("--timeout",  type=int, default=120,
                    help="HTTP timeout per request in seconds (default: 120)")
parser.add_argument("--delay",    type=float, default=0.5,
                    help="Delay between requests in seconds (default: 0.5)")
args = parser.parse_args()

API_URL   = args.api_url.rstrip("/")
MATCH_URL = f"{API_URL}/match"
TIMEOUT   = args.timeout
DELAY     = args.delay

# ── Colours ───────────────────────────────────────────────────────────────────
C_HEADER_DARK  = "1F3864"
C_HEADER_MID   = "2E75B6"
C_MATCH        = "C6EFCE"   # green  – correct match
C_WRONG_POS    = "FFEB9C"   # amber  – right person, wrong rank
C_MISS         = "FFC7CE"   # red    – not in top-3 at all
C_NEUTRAL      = "FFFFFF"
C_ALT_ROW      = "EEF4FB"
C_SUMMARY_HDR  = "2C5F8A"

thin  = Side(style="thin",   color="AAAAAA")
thick = Side(style="medium", color="888888")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ══════════════════════════════════════════════════════════════════════════════
# 1. DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════

def load_data(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, dtype=str).fillna("")
    required = {"userID", "userName", "userprofileinfo",
                "userobjective", "match target 1", "match target 2", "match target 3"}
    missing = required - set(df.columns)
    if missing:
        sys.exit(f"ERROR: Missing columns in input file: {missing}")
    print(f"✅ Loaded {len(df)} users from {path}")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 2. SCHEMA BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

TITLE_KEYWORD_MAP = {
    "trader":       ["Trader", "Equity Trader", "Quantitative Trader"],
    "engineer":     ["Software Engineer", "Data Engineer", "ML Engineer"],
    "data":         ["Data Scientist", "Data Analyst", "ML Engineer"],
    "scientist":    ["Data Scientist", "Research Scientist"],
    "designer":     ["Product Designer", "UX Designer", "Designer"],
    "manager":      ["Product Manager", "Project Manager", "Operations Manager"],
    "partner":      ["Partner", "Managing Partner", "Strategic Partner"],
    "auditor":      ["ISO Auditor", "Compliance Analyst", "Risk Auditor"],
    "analyst":      ["Business Analyst", "Compliance Analyst", "Risk Analyst"],
    "consultant":   ["Consultant", "Advisory Consultant"],
    "lawyer":       ["Lawyer", "Legal Counsel", "Partner"],
    "founder":      ["Co-Founder", "Founder", "CEO"],
    "investor":     ["Investor", "VC", "Angel Investor"],
    "researcher":   ["Researcher", "Research Scientist"],
    "director":     ["Director", "Managing Director"],
}

def _infer_target_profiles(objective_text: str, profile_json: dict) -> List[dict]:
    """
    Infer target_profiles from the objective text.
    Looks for role/title keywords and groups them into TargetProfile dicts.
    """
    obj_lower = objective_text.lower()
    seen_titles = set()
    profiles = []

    for keyword, titles in TITLE_KEYWORD_MAP.items():
        if keyword in obj_lower:
            for t in titles:
                if t not in seen_titles:
                    seen_titles.add(t)
            profiles.append({
                "type": keyword,
                "titles": titles,
                "why": f"Objective mentions '{keyword}'"
            })

    # Fallback: use the user's own role to cast a wide net
    if not profiles:
        own_title = profile_json.get("current_role", {}).get("title", "Professional")
        profiles.append({
            "type": "general",
            "titles": [own_title, "Consultant", "Director", "Manager"],
            "why": "General networking based on own profile"
        })

    return profiles


def _infer_industry(profile_json: dict) -> str:
    """Guess an industry tag from solutions_offered / career_highlights."""
    text = " ".join(
        profile_json.get("solutions_offered", []) +
        profile_json.get("career_highlights", [])
    ).lower()
    for kw, ind in [
        ("cyber",      "Cybersecurity"),
        ("finance",    "Financial Services"),
        ("healthcare", "Healthcare"),
        ("blockchain", "Blockchain / Web3"),
        ("ai",         "Artificial Intelligence"),
        ("ml",         "Machine Learning"),
        ("trading",    "Capital Markets"),
        ("design",     "Product Design"),
        ("legal",      "Legal"),
        ("data",       "Data & Analytics"),
    ]:
        if kw in text:
            return ind
    return "Technology"


def build_user_profile_payload(profile_json: dict) -> dict:
    """Map userprofileinfo JSON → UserProfileInfo API payload."""
    return {
        "current_role":    profile_json.get("current_role", {}),
        "previous_roles":  profile_json.get("previous_roles", []),
        "top_skills":      profile_json.get("top_skills", []),
        "solutions_offered": profile_json.get("solutions_offered", []),
        "career_highlights": profile_json.get("career_highlights", []),
    }


def build_user_objective_payload(
    user_id: str,
    objective_text: str,
    profile_json: dict,
) -> dict:
    """Map semicolon-delimited objective text → UserObjective API payload."""
    sentences = [s.strip() for s in objective_text.split(";") if s.strip()]
    primary   = sentences[0] if sentences else "Connect with relevant professionals"
    secondary = sentences[1:] if len(sentences) > 1 else []

    return {
        "person_id":       user_id,
        "primary_goal":    primary,
        "secondary_goals": secondary,
        "target_profiles": _infer_target_profiles(objective_text, profile_json),
        "exclude":         [],
        "success_signals": sentences,   # full objective as retrieval signals
    }


def build_network_profile(row: pd.Series) -> dict:
    """Map a user row → NetworkProfile API payload (used as a candidate)."""
    try:
        pjson = json.loads(row["userprofileinfo"])
    except Exception:
        pjson = {}

    skills = [sk["skill"] for sk in pjson.get("top_skills", []) if "skill" in sk]

    return {
        "profile_id": row["userID"],
        "name":       row["userName"],
        "title":      pjson.get("current_role", {}).get("title", "Professional"),
        "company":    pjson.get("current_role", {}).get("company"),
        "industry":   _infer_industry(pjson),
        "skills":     skills,
        "summary":    "; ".join(pjson.get("solutions_offered", [])),
    }


def build_match_request(user_row: pd.Series, all_rows: pd.DataFrame) -> dict:
    """Assemble the full MatchRequest payload for a single user."""
    try:
        profile_json = json.loads(user_row["userprofileinfo"])
    except Exception:
        profile_json = {}

    user_id     = user_row["userID"]
    candidates  = [
        build_network_profile(row)
        for _, row in all_rows.iterrows()
        if row["userID"] != user_id
    ]

    return {
        "user_profile":    build_user_profile_payload(profile_json),
        "user_objective":  build_user_objective_payload(
                               user_id, user_row["userobjective"], profile_json),
        "network_profiles": candidates,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 3. API CALLER
# ══════════════════════════════════════════════════════════════════════════════

def check_api_health():
    """Verify the API is reachable before running tests."""
    try:
        r = requests.get(f"{API_URL}/health", timeout=10)
        r.raise_for_status()
        print(f"✅ API reachable at {API_URL}")
    except Exception as e:
        print(f"\n❌ Cannot reach API at {API_URL}")
        print(f"   Error: {e}")
        print(f"\n   Make sure the API is running:")
        print(f"   cd <your-api-folder> && uvicorn main:app --host 0.0.0.0 --port 8000")
        sys.exit(1)


def call_match_api(
    payload: dict,
    user_id: str,
    user_name: str,
) -> Tuple[Optional[List[dict]], Optional[str]]:
    """
    POST to /match.  Returns (results_list, error_message).
    results_list is None on error.
    """
    try:
        r = requests.post(MATCH_URL, json=payload, timeout=TIMEOUT)
        r.raise_for_status()
        return r.json(), None
    except requests.exceptions.Timeout:
        return None, f"Timed out after {TIMEOUT}s"
    except requests.exceptions.HTTPError as e:
        body = ""
        try:
            body = r.json()
        except Exception:
            pass
        return None, f"HTTP {r.status_code}: {body or str(e)}"
    except Exception as e:
        return None, str(e)


# ══════════════════════════════════════════════════════════════════════════════
# 4. COMPARISON LOGIC
# ══════════════════════════════════════════════════════════════════════════════

def compare_matches(
    api_top3:       List[str],   # profile_ids from API, ranked 1-3
    expected_top3:  List[str],   # expected profile_ids from sheet, ranked 1-3
) -> Dict:
    """
    Returns per-rank comparison and aggregate metrics.

    Status per slot:
      EXACT       – correct ID at the same rank position
      WRONG_RANK  – correct ID present but at a different rank
      MISS        – expected ID not returned anywhere in top-3
      EXTRA       – API returned a candidate not in expected set
    """
    api_set      = set(api_top3)
    expected_set = set(expected_top3)

    per_rank = []
    for rank_idx, exp_id in enumerate(expected_top3):
        if rank_idx < len(api_top3) and api_top3[rank_idx] == exp_id:
            status = "EXACT"
        elif exp_id in api_set:
            api_rank = api_top3.index(exp_id) + 1
            status = f"WRONG_RANK (API rank {api_rank})"
        else:
            status = "MISS"
        per_rank.append(status)

    exact_count     = sum(1 for s in per_rank if s == "EXACT")
    in_top3_count   = sum(1 for s in per_rank if "MISS" not in s)
    overlap         = len(api_set & expected_set)

    return {
        "per_rank":    per_rank,
        "exact":       exact_count,
        "in_top3":     in_top3_count,
        "overlap":     overlap,
        "accuracy":    exact_count / max(len(expected_top3), 1),
    }


# ══════════════════════════════════════════════════════════════════════════════
# 5. EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

def _hcell(cell, value, bg=C_HEADER_DARK, bold=True, fc="FFFFFF", size=9, wrap=True):
    cell.value     = value
    cell.font      = Font(name="Arial", bold=bold, color=fc, size=size)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border    = BORDER


def _dcell(cell, value, bg=C_NEUTRAL, bold=False, fc="000000", size=9,
           halign="left", wrap=True):
    cell.value     = value
    cell.font      = Font(name="Arial", bold=bold, color=fc, size=size)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=halign, vertical="top", wrap_text=wrap)
    cell.border    = BORDER


def _status_color(status: str) -> str:
    if status == "EXACT":
        return C_MATCH
    elif "WRONG_RANK" in status:
        return C_WRONG_POS
    else:
        return C_MISS


def build_output_excel(results: List[dict], output_path: str):
    wb = Workbook()

    # ── Sheet 1: Full Detail ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Match Comparison"

    # Row 1: Group headers
    ws.merge_cells("A1:D1"); _hcell(ws["A1"], "USER",            C_HEADER_DARK)
    ws.merge_cells("E1:I1"); _hcell(ws["E1"], "RANK 1",          "1A6B3C")
    ws.merge_cells("J1:N1"); _hcell(ws["J1"], "RANK 2",          "9C6E00")
    ws.merge_cells("O1:S1"); _hcell(ws["O1"], "RANK 3",          "8B2500")
    ws.merge_cells("T1:U1"); _hcell(ws["T1"], "SUMMARY",         C_HEADER_MID)
    ws.row_dimensions[1].height = 22

    # Row 2: Column headers
    col_headers = [
        # User (A-D)
        "User ID", "User Name", "Objective (summary)", "API Status",
        # Rank 1 (E-I)
        "Expected ID", "Expected Name",
        "API ID", "API Name", "Result",
        # Rank 2 (J-N)
        "Expected ID", "Expected Name",
        "API ID", "API Name", "Result",
        # Rank 3 (O-S)
        "Expected ID", "Expected Name",
        "API ID", "API Name", "Result",
        # Summary (T-U)
        "Exact Matches", "Overlap / 3",
    ]
    BG_ROW2 = ["2C5F8A"]*4 + ["1A6B3C"]*5 + ["9C6E00"]*5 + ["8B2500"]*5 + [C_SUMMARY_HDR]*2
    for col_idx, (hdr, bg) in enumerate(zip(col_headers, BG_ROW2), start=1):
        _hcell(ws.cell(row=2, column=col_idx), hdr, bg=bg, size=8)
    ws.row_dimensions[2].height = 32

    # Data rows
    for row_idx, res in enumerate(results, start=3):
        bg_base = C_ALT_ROW if row_idx % 2 == 0 else C_NEUTRAL

        uid       = res["user_id"]
        uname     = res["user_name"]
        obj_short = res["objective_short"]
        api_err   = res.get("error")
        expected  = res["expected"]   # list of {"id", "name"}
        api_res   = res["api_results"] # list of {"id", "name", "score", "reason"}
        cmp       = res.get("comparison", {})

        # User columns (A-D)
        _dcell(ws.cell(row=row_idx, column=1), uid,       bg=bg_base, halign="center")
        _dcell(ws.cell(row=row_idx, column=2), uname,     bg=bg_base, bold=True)
        _dcell(ws.cell(row=row_idx, column=3), obj_short, bg=bg_base)
        status_text = "✅ OK" if not api_err else f"❌ {api_err}"
        status_bg   = C_MATCH if not api_err else C_MISS
        _dcell(ws.cell(row=row_idx, column=4), status_text, bg=status_bg,
               bold=True, halign="center")

        # Rank blocks (E-I, J-N, O-S)
        col_offset = 5
        for rank_i in range(3):
            exp   = expected[rank_i]   if rank_i < len(expected)  else {"id": "", "name": ""}
            apir  = api_res[rank_i]    if rank_i < len(api_res)   else {"id": "", "name": ""}
            per_r = cmp.get("per_rank", ["N/A", "N/A", "N/A"])[rank_i] if cmp else "ERROR"

            res_bg = _status_color(per_r) if not api_err else "D3D3D3"

            _dcell(ws.cell(row=row_idx, column=col_offset+0), exp["id"],   bg=bg_base, halign="center")
            _dcell(ws.cell(row=row_idx, column=col_offset+1), exp["name"], bg=bg_base)
            _dcell(ws.cell(row=row_idx, column=col_offset+2), apir["id"],  bg=bg_base, halign="center")
            _dcell(ws.cell(row=row_idx, column=col_offset+3), apir["name"],bg=bg_base)
            _dcell(ws.cell(row=row_idx, column=col_offset+4), per_r,       bg=res_bg,
                   bold=(per_r == "EXACT"), halign="center")
            col_offset += 5

        # Summary (T-U)
        exact   = cmp.get("exact",   0) if cmp else 0
        overlap = cmp.get("overlap", 0) if cmp else 0
        _dcell(ws.cell(row=row_idx, column=20), f"{exact}/3",
               bg=C_MATCH if exact == 3 else (C_WRONG_POS if exact > 0 else C_MISS),
               bold=True, halign="center")
        _dcell(ws.cell(row=row_idx, column=21), f"{overlap}/3",
               bg=bg_base, halign="center")

        ws.row_dimensions[row_idx].height = 40

    # Column widths
    WIDTHS = [18, 16, 45, 12,   # User
              16, 16, 16, 16, 16,  # R1
              16, 16, 16, 16, 16,  # R2
              16, 16, 16, 16, 16,  # R3
              10, 10]              # Summary
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "E3"

    # ── Sheet 2: API Raw Scores ────────────────────────────────────────────────
    ws2 = wb.create_sheet("API Raw Scores")
    raw_headers = ["User ID", "User Name",
                   "API Rank", "Candidate ID", "Candidate Name",
                   "Score", "KG Signals", "Retrieval Rank", "Reason"]
    for ci, h in enumerate(raw_headers, start=1):
        _hcell(ws2.cell(row=1, column=ci), h, bg=C_HEADER_MID)
    ws2.row_dimensions[1].height = 25

    raw_row = 2
    for res in results:
        if res.get("error"):
            ws2.cell(row=raw_row, column=1, value=res["user_id"])
            ws2.cell(row=raw_row, column=2, value=res["user_name"])
            ws2.cell(row=raw_row, column=3, value=f"ERROR: {res['error']}")
            raw_row += 1
            continue
        bg = C_ALT_ROW if raw_row % 2 == 0 else C_NEUTRAL
        for rank_i, apir in enumerate(res["api_results"], start=1):
            ws2.cell(row=raw_row, column=1, value=res["user_id"])
            ws2.cell(row=raw_row, column=2, value=res["user_name"])
            ws2.cell(row=raw_row, column=3, value=rank_i)
            ws2.cell(row=raw_row, column=4, value=apir.get("id", ""))
            ws2.cell(row=raw_row, column=5, value=apir.get("name", ""))
            ws2.cell(row=raw_row, column=6, value=apir.get("score", ""))
            ws2.cell(row=raw_row, column=7, value="; ".join(apir.get("kg_signals", [])))
            ws2.cell(row=raw_row, column=8, value=apir.get("retrieval_rank", ""))
            ws2.cell(row=raw_row, column=9, value=apir.get("reason", ""))
            for ci in range(1, 10):
                c = ws2.cell(row=raw_row, column=ci)
                c.fill   = PatternFill("solid", start_color=bg)
                c.font   = Font(name="Arial", size=9)
                c.border = BORDER
                c.alignment = Alignment(vertical="top", wrap_text=True)
            raw_row += 1

    for i, w in enumerate([18,16, 8, 18,18, 8, 60, 12, 60], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Summary Stats ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    success = [r for r in results if not r.get("error")]
    errors  = [r for r in results if r.get("error")]

    total         = len(results)
    total_ok      = len(success)
    total_exact3  = sum(1 for r in success if r["comparison"]["exact"] == 3)
    total_exact2  = sum(1 for r in success if r["comparison"]["exact"] == 2)
    total_exact1  = sum(1 for r in success if r["comparison"]["exact"] == 1)
    total_exact0  = sum(1 for r in success if r["comparison"]["exact"] == 0)
    avg_overlap   = (sum(r["comparison"]["overlap"] for r in success) / max(total_ok, 1))
    avg_exact     = (sum(r["comparison"]["exact"]   for r in success) / max(total_ok, 1))

    stats = [
        ("OVERALL ACCURACY", ""),
        ("Users tested",           total),
        ("API calls successful",   total_ok),
        ("API call errors",        len(errors)),
        ("",                       ""),
        ("RANK MATCHING",          ""),
        ("All 3 exact matches",    f"{total_exact3} / {total_ok}"),
        ("Exactly 2 matches",      f"{total_exact2} / {total_ok}"),
        ("Exactly 1 match",        f"{total_exact1} / {total_ok}"),
        ("No matches",             f"{total_exact0} / {total_ok}"),
        ("",                       ""),
        ("AVERAGES (successful calls)", ""),
        ("Avg exact rank matches / user",  f"{avg_exact:.2f} / 3"),
        ("Avg top-3 overlap / user",       f"{avg_overlap:.2f} / 3"),
        ("",                       ""),
        ("LEGEND",                 ""),
        ("EXACT",      "Correct candidate at the correct rank"),
        ("WRONG_RANK", "Correct candidate but at different rank"),
        ("MISS",       "Expected candidate not in API top-3"),
    ]

    legend_bg = {
        "EXACT":       C_MATCH,
        "WRONG_RANK":  C_WRONG_POS,
        "MISS":        C_MISS,
        "OVERALL ACCURACY": C_HEADER_DARK,
        "RANK MATCHING":    C_HEADER_DARK,
        "AVERAGES (successful calls)": C_HEADER_DARK,
        "LEGEND":          C_HEADER_DARK,
    }

    for si, (label, value) in enumerate(stats, start=1):
        c1 = ws3.cell(row=si, column=1, value=label)
        c2 = ws3.cell(row=si, column=2, value=value)
        bg = legend_bg.get(label, C_NEUTRAL)
        for c in (c1, c2):
            c.fill   = PatternFill("solid", start_color=bg)
            c.border = BORDER
            c.font   = Font(name="Arial",
                            bold=(label in legend_bg),
                            color="FFFFFF" if label in legend_bg else "000000",
                            size=10)
            c.alignment = Alignment(vertical="center", horizontal="left")
        ws3.row_dimensions[si].height = 20

    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 30

    wb.save(output_path)
    print(f"\n✅ Results saved to: {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# 6. MAIN RUNNER
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print(f"\n{'─'*60}")
    print(f"  Rain 2.0 API Test Runner")
    print(f"  API:    {API_URL}")
    print(f"  Input:  {args.input}")
    print(f"  Output: {args.output}")
    print(f"{'─'*60}\n")

    # Health check
    check_api_health()

    # Load data
    df = load_data(args.input)
    id_to_name = dict(zip(df["userID"], df["userName"]))

    all_results = []

    for idx, user_row in df.iterrows():
        uid    = user_row["userID"]
        uname  = user_row["userName"]
        obj    = user_row["userobjective"]
        # Short objective for display (first sentence)
        obj_short = obj.split(";")[0].strip()[:120]

        expected_ids = [
            user_row.get("match target 1", "").strip(),
            user_row.get("match target 2", "").strip(),
            user_row.get("match target 3", "").strip(),
        ]
        expected = [
            {"id": eid, "name": id_to_name.get(eid, "Unknown")}
            for eid in expected_ids
        ]

        print(f"[{idx+1:02d}/{len(df)}] {uname:<22} → calling API ...", end=" ", flush=True)

        payload  = build_match_request(user_row, df)
        api_data, err = call_match_api(payload, uid, uname)

        if err:
            print(f"❌ ERROR: {err}")
            all_results.append({
                "user_id": uid, "user_name": uname,
                "objective_short": obj_short,
                "expected":  expected,
                "api_results": [],
                "comparison": None,
                "error": err,
            })
        else:
            # api_data is a list of MatchResult dicts, already sorted by score desc
            api_top3 = api_data[:3]
            api_top3_ids = [r["profile_id"] for r in api_top3]
            api_top3_fmt = [
                {
                    "id":             r["profile_id"],
                    "name":           r["name"],
                    "score":          r["score"],
                    "reason":         r.get("reason", ""),
                    "kg_signals":     r.get("kg_signals", []),
                    "retrieval_rank": r.get("retrieval_rank"),
                }
                for r in api_top3
            ]

            cmp = compare_matches(api_top3_ids, expected_ids)
            emoji = "✅" if cmp["exact"] == 3 else ("🟡" if cmp["exact"] > 0 else "🔴")
            print(f"{emoji}  exact={cmp['exact']}/3  overlap={cmp['overlap']}/3")

            all_results.append({
                "user_id":        uid,
                "user_name":      uname,
                "objective_short": obj_short,
                "expected":       expected,
                "api_results":    api_top3_fmt,
                "comparison":     cmp,
                "error":          None,
            })

        time.sleep(DELAY)

    # Print quick summary
    success = [r for r in all_results if not r.get("error")]
    if success:
        avg_e = sum(r["comparison"]["exact"] for r in success) / len(success)
        avg_o = sum(r["comparison"]["overlap"] for r in success) / len(success)
        print(f"\n{'─'*60}")
        print(f"  DONE  — {len(success)}/{len(all_results)} calls succeeded")
        print(f"  Avg exact rank matches : {avg_e:.2f} / 3")
        print(f"  Avg top-3 overlap      : {avg_o:.2f} / 3")
        print(f"{'─'*60}")

    build_output_excel(all_results, args.output)


if __name__ == "__main__":
    main()
